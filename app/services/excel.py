"""Import/export de produtos em Excel (openpyxl), respeitando campos customizados."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.extensions import db
from app.models.categoria import Categoria
from app.models.definicao_campo import ENTIDADE_PRODUTO
from app.models.produto import TIPOS_CONTROLE, Produto
from app.services import campos_customizados as cc
from app.services import produto_service
from app.services.produto_service import ErroProduto

# Colunas fixas exportadas/importadas.
COLUNAS_FIXAS = [
    ("sku", "SKU"),
    ("nome", "Nome"),
    ("tipo_controle", "Tipo (CONSUMIVEL/DURAVEL)"),
    ("categoria", "Categoria"),
    ("unidade", "Unidade"),
    ("estoque_minimo", "Estoque mínimo"),
    ("estoque_maximo", "Estoque máximo"),
    ("marca", "Marca"),
    ("modelo", "Modelo"),
    ("valor_unitario_referencia", "Valor ref. (R$)"),
    ("descricao", "Descrição"),
]


def exportar_produtos(organizacao_id: int, *, incluir_inativos: bool = False) -> bytes:
    """Gera uma planilha .xlsx com os produtos da organização."""
    from sqlalchemy import select

    # Definições (globais + de qualquer categoria) como colunas adicionais.
    from app.models.definicao_campo import DefinicaoCampo

    definicoes = list(
        db.session.scalars(
            select(DefinicaoCampo)
            .where(
                DefinicaoCampo.organizacao_id == organizacao_id,
                DefinicaoCampo.entidade == ENTIDADE_PRODUTO,
                DefinicaoCampo.ativo.is_(True),
            )
            .order_by(DefinicaoCampo.ordem, DefinicaoCampo.rotulo)
        )
    )

    stmt = select(Produto).where(Produto.organizacao_id == organizacao_id)
    if not incluir_inativos:
        stmt = stmt.where(Produto.ativo.is_(True))
    produtos = list(db.session.scalars(stmt.order_by(Produto.nome)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"

    cabecalhos = [rotulo for _, rotulo in COLUNAS_FIXAS] + [
        f"[{d.chave}] {d.rotulo}" for d in definicoes
    ]
    ws.append(cabecalhos)
    for col in range(1, len(cabecalhos) + 1):
        cel = ws.cell(row=1, column=col)
        cel.font = Font(bold=True, color="FFFFFF")
        cel.fill = PatternFill("solid", fgColor="0D6EFD")
        ws.column_dimensions[get_column_letter(col)].width = 20

    for p in produtos:
        linha: list[Any] = [
            p.sku,
            p.nome,
            p.tipo_controle,
            p.categoria.nome if p.categoria else "",
            p.unidade,
            float(p.estoque_minimo or 0),
            float(p.estoque_maximo) if p.estoque_maximo is not None else "",
            p.marca or "",
            p.modelo or "",
            float(p.valor_unitario_referencia) if p.valor_unitario_referencia is not None else "",
            p.descricao or "",
        ]
        for d in definicoes:
            linha.append(cc.formatar_valor(d, p.campos.get(d.chave)))
        ws.append(linha)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def gerar_modelo_importacao() -> bytes:
    """Planilha vazia (só cabeçalhos das colunas fixas) para o usuário preencher."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"
    ws.append([rotulo for _, rotulo in COLUNAS_FIXAS])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def importar_produtos(organizacao_id: int, conteudo: bytes) -> dict[str, Any]:
    """Importa produtos da planilha. Retorna resumo com criados/erros.

    Estratégia: valida todas as linhas; cria as válidas. Linhas com SKU existente
    são atualizadas (nome/categoria/estoques/valores). Erros não impedem as demais.
    """
    from sqlalchemy import select

    wb = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb.active

    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        return {"criados": 0, "atualizados": 0, "erros": [{"linha": 0, "msg": "Planilha vazia."}]}

    categorias = {
        c.nome.lower(): c
        for c in db.session.scalars(
            select(Categoria).where(Categoria.organizacao_id == organizacao_id)
        )
    }

    criados = atualizados = 0
    erros: list[dict] = []

    for n, linha in enumerate(linhas[1:], start=2):  # pula cabeçalho
        valores = list(linha) + [None] * (len(COLUNAS_FIXAS) - len(linha))
        sku = str(valores[0]).strip() if valores[0] else ""
        nome = str(valores[1]).strip() if valores[1] else ""
        if not nome:
            if any(v not in (None, "") for v in valores):
                erros.append({"linha": n, "msg": "Nome é obrigatório."})
            continue

        tipo = str(valores[2]).strip().upper() if valores[2] else "CONSUMIVEL"
        if tipo not in TIPOS_CONTROLE:
            erros.append({"linha": n, "msg": f"Tipo inválido: {tipo}"})
            continue

        cat_nome = str(valores[3]).strip() if valores[3] else ""
        categoria_id = categorias[cat_nome.lower()].id if cat_nome.lower() in categorias else None

        dados = {
            "nome": nome,
            "tipo_controle": tipo,
            "categoria_id": categoria_id,
            "unidade": (str(valores[4]).strip().upper() if valores[4] else "UN"),
            "estoque_minimo": _num(valores[5]) or 0,
            "estoque_maximo": _num(valores[6]),
            "marca": (str(valores[7]).strip() if valores[7] else None),
            "modelo": (str(valores[8]).strip() if valores[8] else None),
            "valor_unitario_referencia": _num(valores[9]),
            "descricao": (str(valores[10]).strip() if valores[10] else None),
        }

        try:
            existente = None
            if sku:
                existente = db.session.scalar(
                    select(Produto).where(
                        Produto.organizacao_id == organizacao_id, Produto.sku == sku
                    )
                )
            if existente:
                produto_service.atualizar_produto(existente, dados=dados, commit=False)
                atualizados += 1
            else:
                produto_service.criar_produto(
                    organizacao_id, sku=sku or None, **dados, commit=False
                )
                criados += 1
        except ErroProduto as exc:
            erros.append({"linha": n, "msg": str(exc)})

    if erros and not (criados or atualizados):
        db.session.rollback()
    else:
        db.session.commit()
    return {"criados": criados, "atualizados": atualizados, "erros": erros}


def _num(valor: Any) -> float | None:
    if valor in (None, ""):
        return None
    try:
        if isinstance(valor, str):
            valor = valor.replace(".", "").replace(",", ".") if "," in valor else valor
        return float(valor)
    except (ValueError, TypeError):
        return None
